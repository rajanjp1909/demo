#create a Excel file
import openpyxl, os, time
print(dir(openpyxl))

from openpyxl import Workbook
wb = Workbook()
print(wb)
sheet = wb.active
print(sheet)
sheet["A1"] = "John"
sheet["B1"] = "Trainee"
wb.save(filename = "new sheet.xlsx")
print(os.getcwd())

sheet1 = wb.create_sheet("first")
sheet2 = wb.create_sheet("second")
sheet3 = wb.create_sheet()
wb.save(filename = "new sheet.xlsx")

sheet4 = wb.create_sheet("third", 0)
sheet5 = wb.create_sheet("fourth", 1)
wb.save(filename = "new sheet.xlsx")
sheet3.title = "fifth"
print(wb.save(filename = "new sheet.xlsx"))

sheet["A2"] = 12.5
sheet["A3"] = "Welcome"
now = time.strftime("%x")
sheet["A4"] = now
v1 = sheet5["A6"]
v1.value = 210
wb.save(filename = "new sheet.xlsx")

#change font in excel file
from openpyxl import styles
sheet3["A1"] = "Times new roman"
wb.save(filename = "new sheet.xlsx")
font = styles.Font(name = "Bauhaus 93", size = 40, bold = True, italic = True)
sheet3["A1"].font = font
wb.save(filename = "new sheet.xlsx")

#reading excel files
path = "students.xlsx"
op = openpyxl.load_workbook(path)
sheet = op.active
v1 = sheet.cell(row = 2, column = 1)
print(v1.value)
print(sheet.max_row)
print(sheet.max_column)
max_col = sheet.max_column
for i in range(1, max_col + 1):
    cell_val = sheet.cell(row = 1, column = i)
    print(cell_val.value)
max_row = sheet.max_row
for i in range(1, max_row + 1):
    cell_val = sheet.cell(row = i, column = 1)
    print(cell_val.value)

#merge and unmerge cells
wb = Workbook()
sheet = wb.active
sheet.merge_cells("A1:F1")
wb.save(filename = "Merged ones.xlsx")
sheet.unmerge_cells("A1:F1")
wb.save(filename = "Merged ones.xlsx")
sheet["A1"] = "this is how you will extend the cell"
wb.save(filename = "Merged ones.xlsx")

#adjust length and width of a cell
sheet.row_dimensions[1].height = 70
sheet.column_dimensions["A"].width = 30
wb.save(filename = "Merged ones.xlsx")


