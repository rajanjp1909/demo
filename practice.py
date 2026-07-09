# path = "newfile.txt"
# count = dict()
#
# with open(path, "r") as file:
#     content = file.read()
#     line = content.strip()
#     line = content.lower()
#
#     words= line.split()
#
#     for word in words:
#         if word in count:
#             count[word] += 1
#         else:
#             count[word] = 1
#
#     print(count)

#check bank balance of customers
import openpyxl

class BankAccount:
    def __init__(self, account, name, balance, deposit, debit):
        self.account = account
        self.name = name
        self.balance = balance
        self.deposit = deposit
        self.debit = debit
        self.current_bal = (self.balance + self.deposit) - self.debit

    def make_deposit(self):
        return self.current_bal

    def withdraw(self, amount):
        if amount <= self.current_bal:
            self.current_bal -= amount
            self.balance = self.current_bal
            self.debit = amount
        else:
            print("Not enough money")

    def display(self):
        return {
            "Account": self.account,
            "Name": self.name,
            "Balance": self.balance,
            "Last Deposit": self.deposit,
            "Last debit": self.debit,
            "Current Balance": self.current_bal
        }

path = "accounts.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

accounts = {}
for row in ws.iter_rows(min_row = 2, values_only = True):
    account, name, balance, deposit, debit = row
    b1 = BankAccount(account, name, balance, deposit, debit)
    accounts[account] = b1.display()

y = list(accounts.values())
print(y)
a = input("Enter account no.: ")
if a in accounts:
    print("Customer Details: ", accounts[a])
else:
    print("No data found")

# accounts = []
# for row in ws.iter_rows(min_row = 2, values_only = True):
#     account, name, balance, deposit, debit = row
#     b1 = BankAccount(account, name, balance, deposit, debit)
#     accounts.append(b1.display())
# print(accounts)
#








